"""Import memes from Excel file to database."""

import os
import sys
import logging
from pathlib import Path
from typing import List, Dict

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import select

from db.connection import SessionLocal, init_db
from db.models import Meme

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# OpenAI configuration
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")


def parse_meme_id(value) -> str:
    """
    Parse meme ID from Excel cell value.
    Handles formats like SS0001, SS0002, etc.

    Args:
        value: Cell value from Excel

    Returns:
        Meme ID string
    """
    if value is None:
        return ""
    return str(value).strip().upper()


def parse_name(value) -> str:
    """
    Parse meme name from Excel cell value.

    Args:
        value: Cell value from Excel

    Returns:
        Name string
    """
    if value is None:
        return ""
    return str(value).strip()


def parse_aliases(value) -> list[str]:
    """
    Parse aliases from Excel cell value.

    Args:
        value: Cell value from Excel (comma-separated)

    Returns:
        List of alias strings
    """
    if value is None:
        return []
    aliases_str = str(value).strip()
    if not aliases_str:
        return []
    return [alias.strip() for alias in aliases_str.split(",") if alias.strip()]


def generate_aliases_with_openai(
    names: List[str], batch_size: int = 50
) -> Dict[str, List[str]]:
    """
    Generate aliases for meme names using OpenAI API (batch processing).

    Args:
        names: List of meme names
        batch_size: Number of names to process in each batch

    Returns:
        Dictionary mapping name to list of aliases
    """
    if not OPENAI_API_KEY:
        logger.warning("OPENAI_API_KEY not set, skipping alias generation")
        return {}

    client = OpenAI(api_key=OPENAI_API_KEY)
    results = {}

    # Process in batches
    for i in range(0, len(names), batch_size):
        batch = names[i : i + batch_size]
        logger.info(
            f"Generating aliases for batch {i//batch_size + 1} ({len(batch)} memes)..."
        )

        try:
            # Create prompt for batch
            names_list = "\n".join([f"- {name}" for name in batch])
            prompt = f"""
請為以下「海綿寶寶梗圖名稱」生成搜尋用 aliases。

需求說明：
- 每個梗圖名稱請提供 2–4 個 aliases
- 請嚴格遵守 system prompt 中的所有規則

梗圖名稱列表：
{names_list}

請只回傳 JSON，不要任何解釋或多餘文字。
輸出格式必須完全符合以下結構：

{{
  "梗圖名稱1": ["alias1", "alias2", "alias3"],
  "梗圖名稱2": ["alias1", "alias2", "alias3"]
}}
"""

            response = client.chat.completions.create(
                model=OPENAI_MODEL,
                messages=[
                    {
                        "role": "system",
                        "content": """
你是搜尋關鍵字設計專家，負責為梗圖系統產生 aliases。

aliases 將用於 PostgreSQL pg_trgm 模糊搜尋（短字串比對）。
請遵守：
- alias 長度 2–4 字（最多 5）
- 取句子部分片段、關鍵字

輸出必須是合法 JSON，將被程式直接解析。
""",
                    },
                    {"role": "user", "content": prompt},
                ],
                response_format={"type": "json_object"},
            )

            import json

            content = response.choices[0].message.content
            if not content:
                raise ValueError("Empty response from OpenAI")
            aliases_dict = json.loads(content)

            # Update results
            for name in batch:
                if name in aliases_dict:
                    results[name] = aliases_dict[name]
                else:
                    logger.warning(f"No aliases generated for: {name}")
                    results[name] = []

        except Exception as e:
            logger.error(f"Error generating aliases for batch: {e}")
            # Fallback: set empty aliases for this batch
            for name in batch:
                results[name] = []

    return results


def read_xlsx(file_path: Path) -> List[Dict]:
    """
    Read memes from Excel file.

    Expected format:
    - Column A: ID (e.g., SS0001, SS0002)
    - Column B: Name
    - Column C: Aliases (comma-separated, optional, will be generated if empty)

    Args:
        file_path: Path to Excel file

    Returns:
        List of meme dictionaries
    """
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    memes = []
    # Skip header row (row 1), start from row 2
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not any(row):  # Skip empty rows
            continue

        meme_id = parse_meme_id(row[0])
        name = parse_name(row[1])
        aliases = parse_aliases(row[2]) if len(row) > 2 else []

        if not meme_id or not name:
            logger.warning(f"Row {row_idx}: Missing required fields, skipping")
            continue

        memes.append({"id": meme_id, "name": name, "aliases": aliases})

    return memes


def import_memes_to_db(
    memes: List[Dict], db: Session, update_existing: bool = False
) -> Dict:
    """
    Import memes to database.

    Args:
        memes: List of meme dictionaries
        db: Database session
        update_existing: If True, update existing memes; if False, skip them

    Returns:
        Dictionary with import statistics
    """
    stats = {
        "inserted": 0,
        "updated": 0,
        "skipped": 0,
        "errors": 0,
        "aliases_generated": 0,
    }

    # First, check which memes will be skipped (if update_existing=False)
    # We need to know this before generating aliases to avoid wasting API calls
    existing_meme_ids = set()
    if not update_existing:
        for meme_data in memes:
            meme_id = meme_data.get("id")
            if meme_id:
                existing = db.execute(
                    select(Meme).where(Meme.meme_id == meme_id)
                ).scalar_one_or_none()
                if existing:
                    existing_meme_ids.add(meme_id)

    # Find memes that need aliases and won't be skipped
    memes_needing_aliases = [
        m
        for m in memes
        if not m.get("aliases") and m.get("id") not in existing_meme_ids
    ]
    if memes_needing_aliases:
        logger.info(
            f"Generating aliases for {len(memes_needing_aliases)} memes using OpenAI..."
        )
        names_to_generate = [m["name"] for m in memes_needing_aliases]
        generated_aliases = generate_aliases_with_openai(names_to_generate)

        # Update memes with generated aliases
        for meme in memes_needing_aliases:
            if meme["name"] in generated_aliases:
                meme["aliases"] = generated_aliases[meme["name"]]
                stats["aliases_generated"] += 1

    # Import to database
    for meme_data in memes:
        try:
            meme_id = meme_data["id"]
            name = meme_data["name"]
            aliases = meme_data.get("aliases", [])

            # Check if meme exists
            existing = db.execute(
                select(Meme).where(Meme.meme_id == meme_id)
            ).scalar_one_or_none()

            if existing:
                if update_existing:
                    existing.name = name
                    existing.aliases = aliases
                    stats["updated"] += 1
                    logger.info(f"Updated meme: {meme_id}")
                else:
                    stats["skipped"] += 1
                    logger.debug(f"Skipped existing meme: {meme_id}")
            else:
                # Create new meme
                new_meme = Meme(meme_id=meme_id, name=name, aliases=aliases)
                db.add(new_meme)
                stats["inserted"] += 1
                logger.info(f"Inserted meme: {meme_id} - {name}")

        except Exception as e:
            stats["errors"] += 1
            logger.error(f"Error importing meme {meme_data.get('id', 'unknown')}: {e}")

    # Commit all changes
    try:
        db.commit()
        logger.info("All changes committed to database")
    except Exception as e:
        db.rollback()
        logger.error(f"Error committing to database: {e}")
        raise

    return stats


def main():
    """Main function for importing Excel file."""
    if len(sys.argv) < 2:
        print("Usage: python tools/import_xlsx.py <excel_file> [--update]")
        print("\nExcel format:")
        print("  Column A: ID (e.g., SS0001, SS0002)")
        print("  Column B: Name")
        print("  Column C: Aliases (comma-separated, optional)")
        print("            If empty, aliases will be generated using OpenAI API")
        print("\nOptions:")
        print("  --update: Update existing memes instead of skipping them")
        print(
            "\nNote: Images should be uploaded to R2 at spongebob-memes/{meme_id}.jpg"
        )
        print("\nEnvironment variables:")
        print("  OPENAI_API_KEY: OpenAI API key (required for alias generation)")
        print("  OPENAI_MODEL: OpenAI model to use (default: gpt-4o-mini)")
        sys.exit(1)

    excel_file = Path(sys.argv[1])
    update_existing = "--update" in sys.argv

    if not excel_file.exists():
        print(f"Error: File not found: {excel_file}")
        sys.exit(1)

    # Initialize database
    print("Initializing database...")
    init_db()

    # Read Excel file
    print(f"Reading Excel file: {excel_file}")
    memes = read_xlsx(excel_file)
    print(f"Found {len(memes)} memes in Excel file")

    if not memes:
        print("No memes found in Excel file")
        sys.exit(1)

    # Import to database
    print(f"Importing to database (update_existing={update_existing})...")
    db = SessionLocal()
    try:
        stats = import_memes_to_db(memes, db, update_existing=update_existing)
        print("\nImport completed!")
        print(f"  Inserted: {stats['inserted']}")
        print(f"  Updated: {stats['updated']}")
        print(f"  Skipped: {stats['skipped']}")
        print(f"  Errors: {stats['errors']}")
        if stats["aliases_generated"] > 0:
            print(f"  Aliases generated: {stats['aliases_generated']}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
